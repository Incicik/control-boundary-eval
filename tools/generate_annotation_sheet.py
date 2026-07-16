"""
Generate a human-review annotation sheet from Control-Boundary Eval results.

The utility accepts either:
1. A single evaluation-result JSON file, or
2. A directory containing multiple evaluation-result JSON files.

It produces CSV, XLSX, or both. Existing reviewer annotations can be
preserved when the output CSV already exists.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ANNOTATION_COLUMNS = [
    "annotation_id",
    "scenario_id",
    "category",
    "condition",
    "turn_number",
    "security_boundary",
    "prompt",
    "response",
    "reviewer_1_score",
    "reviewer_1_rationale",
    "evidence_excerpt",
    "refusal",
    "refusal_echo",
    "uncertain",
    "candidate_score_low",
    "candidate_score_high",
    "reviewer_2_score",
    "reviewer_2_rationale",
    "adjudicated_score",
    "adjudication_notes",
    "annotation_status",
    "model_identifier",
    "run_date",
    "run_id",
    "result_file",
]

REVIEW_COLUMNS = [
    "reviewer_1_score",
    "reviewer_1_rationale",
    "evidence_excerpt",
    "refusal",
    "refusal_echo",
    "uncertain",
    "candidate_score_low",
    "candidate_score_high",
    "reviewer_2_score",
    "reviewer_2_rationale",
    "adjudicated_score",
    "adjudication_notes",
]


def parse_arguments() -> argparse.Namespace:
    """Build and parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description=(
            "Convert evaluation-result JSON files into a human-review "
            "annotation sheet."
        )
    )

    parser.add_argument(
        "--input",
        required=True,
        help="Path to one result JSON file or a directory of JSON files.",
    )

    parser.add_argument(
        "--output",
        default="annotations/manual_scores",
        help=(
            "Base output path without an extension. "
            "Default: annotations/manual_scores"
        ),
    )

    parser.add_argument(
        "--format",
        choices=["csv", "xlsx", "both"],
        default="both",
        help="Output format. Default: both",
    )

    parser.add_argument(
        "--preserve",
        action="store_true",
        help=(
            "Preserve reviewer scores and notes from an existing output CSV "
            "when annotation IDs match."
        ),
    )

    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Ask before resolving conflicting duplicate rows.",
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Permit replacement of existing output files.",
    )

    return parser.parse_args()


def discover_result_files(input_path: Path) -> list[Path]:
    """Return one JSON file or all JSON files beneath a directory."""
    if not input_path.exists():
        raise FileNotFoundError(f"Input path does not exist: {input_path}")

    if input_path.is_file():
        if input_path.suffix.lower() != ".json":
            raise ValueError(
                f"Input file must have a .json extension: {input_path}"
            )
        return [input_path]

    result_files = sorted(
        path
        for path in input_path.rglob("*.json")
        if path.is_file()
    )

    if not result_files:
        raise FileNotFoundError(
            f"No JSON result files were found in: {input_path}"
        )

    return result_files


def load_json_file(file_path: Path) -> Any:
    """Load and return JSON content from one file."""
    try:
        with file_path.open("r", encoding="utf-8") as file:
            return json.load(file)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"Malformed JSON in {file_path}: {exc}"
        ) from exc


def normalize_result_records(
    data: Any,
    file_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """
    Normalize supported result-file structures.

    Supported forms:
    - A list of scenario-result records.
    - A dictionary containing a 'results' list plus optional run metadata.
    """
    metadata = {
        "model_identifier": "",
        "run_date": "",
        "run_id": "",
    }

    if isinstance(data, list):
        return data, metadata

    if not isinstance(data, dict):
        raise ValueError(
            f"Unsupported JSON structure in {file_path}. "
            "Expected a list or dictionary."
        )

    metadata["model_identifier"] = str(
        data.get("model_identifier")
        or data.get("model")
        or ""
    )

    metadata["run_date"] = str(
        data.get("run_date")
        or data.get("created_at")
        or data.get("timestamp")
        or ""
    )

    metadata["run_id"] = str(data.get("run_id") or "")

    results = data.get("results")

    if not isinstance(results, list):
        raise ValueError(
            f"Dictionary-form result file {file_path} must contain "
            "a list under the 'results' key."
        )

    return results, metadata


def build_annotation_id(
    scenario_id: str,
    condition: str,
    turn_number: int,
    model_identifier: str,
    run_id: str,
    response: str,
) -> str:
    """
    Create a deterministic annotation ID.

    The same response from the same scenario, condition, model, and run
    will receive the same ID whenever the annotation sheet is regenerated.
    """
    identity = "|".join(
        [
            scenario_id.strip(),
            condition.strip(),
            str(turn_number),
            model_identifier.strip(),
            run_id.strip(),
            response.strip(),
        ]
    )

    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:12].upper()
    return f"ANN-{digest}"


def blank_review_fields() -> dict[str, str]:
    """Return empty reviewer and adjudication fields."""
    return {column: "" for column in REVIEW_COLUMNS}


def create_annotation_row(
    *,
    scenario_id: str,
    category: str,
    condition: str,
    turn_number: int,
    security_boundary: str,
    prompt: str,
    response: str,
    model_identifier: str,
    run_date: str,
    run_id: str,
    result_file: str,
) -> dict[str, Any]:
    """Create one normalized annotation row."""
    annotation_id = build_annotation_id(
        scenario_id=scenario_id,
        condition=condition,
        turn_number=turn_number,
        model_identifier=model_identifier,
        run_id=run_id,
        response=response,
    )

    row: dict[str, Any] = {
        "annotation_id": annotation_id,
        "scenario_id": scenario_id,
        "category": category,
        "condition": condition,
        "turn_number": turn_number,
        "security_boundary": security_boundary,
        "prompt": prompt,
        "response": response,
        **blank_review_fields(),
        "annotation_status": "new",
        "model_identifier": model_identifier,
        "run_date": run_date,
        "run_id": run_id,
        "result_file": result_file,
    }

    return row


def extract_rows_from_record(
    record: dict[str, Any],
    file_path: Path,
    metadata: dict[str, str],
) -> list[dict[str, Any]]:
    """Extract direct and multi-turn annotation rows from one scenario record."""
    scenario_id = str(
        record.get("scenario_id")
        or record.get("id")
        or ""
    ).strip()

    if not scenario_id:
        raise ValueError(
            f"Result record in {file_path} is missing a scenario ID."
        )

    category = str(record.get("category") or "unknown")
    security_boundary = str(record.get("security_boundary") or "")

    model_identifier = str(
        record.get("model_identifier")
        or record.get("model")
        or metadata.get("model_identifier")
        or ""
    )

    run_date = str(
        record.get("run_date")
        or metadata.get("run_date")
        or ""
    )

    run_id = str(
        record.get("run_id")
        or metadata.get("run_id")
        or file_path.stem
    )

    rows: list[dict[str, Any]] = []

    direct_test = record.get("direct_prompt_test", {})
    if isinstance(direct_test, dict):
        direct_prompt = str(direct_test.get("prompt") or "")
        direct_response = str(direct_test.get("response") or "")

        if direct_prompt or direct_response:
            rows.append(
                create_annotation_row(
                    scenario_id=scenario_id,
                    category=category,
                    condition="direct",
                    turn_number=1,
                    security_boundary=security_boundary,
                    prompt=direct_prompt,
                    response=direct_response,
                    model_identifier=model_identifier,
                    run_date=run_date,
                    run_id=run_id,
                    result_file=str(file_path),
                )
            )

    multi_test = record.get("multi_step_prompt_test", {})
    if isinstance(multi_test, dict):
        prompts = multi_test.get("prompts") or []
        responses = multi_test.get("all_responses") or []

        if not isinstance(prompts, list):
            prompts = []

        if not isinstance(responses, list):
            responses = []

        max_turns = max(len(prompts), len(responses))

        for index in range(max_turns):
            prompt = str(prompts[index]) if index < len(prompts) else ""
            response = (
                str(responses[index]) if index < len(responses) else ""
            )

            rows.append(
                create_annotation_row(
                    scenario_id=scenario_id,
                    category=category,
                    condition="multi_turn",
                    turn_number=index + 1,
                    security_boundary=security_boundary,
                    prompt=prompt,
                    response=response,
                    model_identifier=model_identifier,
                    run_date=run_date,
                    run_id=run_id,
                    result_file=str(file_path),
                )
            )

    return rows


def load_existing_annotations(csv_path: Path) -> dict[str, dict[str, str]]:
    """Load existing reviewer fields indexed by stable annotation ID."""
    if not csv_path.exists():
        return {}

    existing: dict[str, dict[str, str]] = {}

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)

        if "annotation_id" not in (reader.fieldnames or []):
            raise ValueError(
                f"Existing CSV lacks annotation_id: {csv_path}"
            )

        for row in reader:
            annotation_id = str(row.get("annotation_id") or "").strip()
            if annotation_id:
                existing[annotation_id] = row

    return existing


def preserve_review_fields(
    row: dict[str, Any],
    existing_row: dict[str, str],
) -> None:
    """Copy reviewer fields from a previously saved annotation row."""
    for column in REVIEW_COLUMNS:
        row[column] = existing_row.get(column, "")

    row["annotation_status"] = "preserved"


def row_source_signature(row: dict[str, Any]) -> tuple[str, ...]:
    """Return source fields used to identify conflicting duplicate rows."""
    return (
        str(row.get("scenario_id", "")),
        str(row.get("condition", "")),
        str(row.get("turn_number", "")),
        str(row.get("prompt", "")),
        str(row.get("response", "")),
        str(row.get("model_identifier", "")),
        str(row.get("run_id", "")),
    )


def resolve_duplicate(
    existing_row: dict[str, Any],
    incoming_row: dict[str, Any],
    interactive: bool,
) -> tuple[dict[str, Any], str]:
    """
    Resolve duplicate annotation IDs.

    Returns:
        selected row and one of:
        - duplicate
        - conflict
    """
    if row_source_signature(existing_row) == row_source_signature(incoming_row):
        return existing_row, "duplicate"

    existing_row["annotation_status"] = "conflict"

    if not interactive:
        return existing_row, "conflict"

    print("\nConflicting annotation rows detected:")
    print(f"Annotation ID: {existing_row['annotation_id']}")
    print(f"Existing result file: {existing_row['result_file']}")
    print(f"Incoming result file: {incoming_row['result_file']}")

    answer = input(
        "Keep [e]xisting, use [i]ncoming, or [a]bort? "
    ).strip().lower()

    if answer == "i":
        incoming_row["annotation_status"] = "conflict"
        return incoming_row, "conflict"

    if answer == "a":
        raise RuntimeError("Aborted due to duplicate conflict.")

    return existing_row, "conflict"


def collect_annotation_rows(
    result_files: list[Path],
    existing_annotations: dict[str, dict[str, str]],
    preserve: bool,
    interactive: bool,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Process result files and return normalized annotation rows."""
    rows_by_id: dict[str, dict[str, Any]] = {}

    summary = {
        "processed_files": 0,
        "created_rows": 0,
        "preserved_rows": 0,
        "duplicate_rows": 0,
        "conflict_rows": 0,
        "skipped_files": 0,
    }

    for file_path in result_files:
        try:
            raw_data = load_json_file(file_path)
            records, metadata = normalize_result_records(
                raw_data,
                file_path,
            )
        except (ValueError, OSError) as exc:
            print(f"[WARNING] Skipping {file_path}: {exc}")
            summary["skipped_files"] += 1
            continue

        summary["processed_files"] += 1

        for record in records:
            if not isinstance(record, dict):
                print(
                    f"[WARNING] Skipping non-dictionary record "
                    f"in {file_path}"
                )
                continue

            try:
                extracted_rows = extract_rows_from_record(
                    record,
                    file_path,
                    metadata,
                )
            except ValueError as exc:
                print(f"[WARNING] {exc}")
                continue

            for row in extracted_rows:
                annotation_id = row["annotation_id"]

                if preserve and annotation_id in existing_annotations:
                    preserve_review_fields(
                        row,
                        existing_annotations[annotation_id],
                    )
                    summary["preserved_rows"] += 1

                if annotation_id in rows_by_id:
                    selected_row, outcome = resolve_duplicate(
                        rows_by_id[annotation_id],
                        row,
                        interactive,
                    )
                    rows_by_id[annotation_id] = selected_row

                    if outcome == "duplicate":
                        summary["duplicate_rows"] += 1
                    else:
                        summary["conflict_rows"] += 1
                else:
                    rows_by_id[annotation_id] = row
                    summary["created_rows"] += 1

    sorted_rows = sorted(
        rows_by_id.values(),
        key=lambda row: (
            str(row["scenario_id"]),
            str(row["condition"]),
            int(row["turn_number"]),
            str(row["model_identifier"]),
            str(row["run_id"]),
        ),
    )

    return sorted_rows, summary


def ensure_output_is_safe(
    output_paths: list[Path],
    overwrite: bool,
    preserve: bool,
) -> None:
    """Prevent accidental replacement of existing annotation work."""
    existing_paths = [path for path in output_paths if path.exists()]

    if not existing_paths:
        return

    if overwrite or preserve:
        return

    joined_paths = "\n".join(f"  - {path}" for path in existing_paths)

    raise FileExistsError(
        "Output file already exists. Use --preserve to retain reviewer "
        "fields or --overwrite to replace it:\n"
        f"{joined_paths}"
    )


def write_csv(rows: list[dict[str, Any]], output_path: Path) -> None:
    """Write the annotation sheet as UTF-8 CSV."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    dataframe = pd.DataFrame(rows, columns=ANNOTATION_COLUMNS)
    dataframe.to_csv(output_path, index=False, encoding="utf-8-sig")


def style_excel(output_path: Path) -> None:
    """Apply basic human-review formatting to the XLSX workbook."""
    from openpyxl import load_workbook

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    worksheet.title = "Manual Scores"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    review_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    status_fill = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9",
    )

    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    header_map = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value
    }

    for column_name in REVIEW_COLUMNS:
        column_index = header_map.get(column_name)
        if not column_index:
            continue

        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_number, column=column_index).fill = (
                review_fill
            )

    status_column = header_map.get("annotation_status")
    if status_column:
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_number,
                column=status_column,
            ).fill = status_fill

    preferred_widths = {
        "annotation_id": 18,
        "scenario_id": 14,
        "category": 28,
        "condition": 14,
        "turn_number": 12,
        "security_boundary": 45,
        "prompt": 55,
        "response": 75,
        "reviewer_1_score": 16,
        "reviewer_1_rationale": 55,
        "evidence_excerpt": 55,
        "reviewer_2_rationale": 55,
        "adjudication_notes": 55,
        "annotation_status": 18,
        "model_identifier": 28,
        "run_date": 18,
        "run_id": 25,
        "result_file": 45,
    }

    for column_name, width in preferred_widths.items():
        column_index = header_map.get(column_name)
        if column_index:
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    workbook.save(output_path)


def write_xlsx(rows: list[dict[str, Any]], output_path: Path) -> None:
    """Write and format the annotation sheet as XLSX."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    dataframe = pd.DataFrame(rows, columns=ANNOTATION_COLUMNS)
    dataframe.to_excel(
        output_path,
        index=False,
        engine="openpyxl",
    )

    style_excel(output_path)


def print_summary(
    summary: dict[str, int],
    output_paths: list[Path],
) -> None:
    """Print a concise completion summary."""
    print("\nAnnotation-sheet generation complete.")
    print(f"Processed result files: {summary['processed_files']}")
    print(f"Skipped result files: {summary['skipped_files']}")
    print(f"Created annotation rows: {summary['created_rows']}")
    print(f"Preserved reviewed rows: {summary['preserved_rows']}")
    print(f"Skipped identical duplicates: {summary['duplicate_rows']}")
    print(f"Conflicts requiring review: {summary['conflict_rows']}")

    print("Output files:")
    for path in output_paths:
        print(f"  - {path}")


def main() -> int:
    """Run the annotation-sheet generator."""
    args = parse_arguments()

    input_path = Path(args.input).expanduser().resolve()
    output_base = Path(args.output).expanduser()

    if output_base.suffix.lower() in {".csv", ".xlsx"}:
        output_base = output_base.with_suffix("")

    csv_path = output_base.with_suffix(".csv")
    xlsx_path = output_base.with_suffix(".xlsx")

    output_paths: list[Path] = []

    if args.format in {"csv", "both"}:
        output_paths.append(csv_path)

    if args.format in {"xlsx", "both"}:
        output_paths.append(xlsx_path)

    try:
        ensure_output_is_safe(
            output_paths,
            overwrite=args.overwrite,
            preserve=args.preserve,
        )

        result_files = discover_result_files(input_path)

        existing_annotations = (
            load_existing_annotations(csv_path)
            if args.preserve
            else {}
        )

        rows, summary = collect_annotation_rows(
            result_files=result_files,
            existing_annotations=existing_annotations,
            preserve=args.preserve,
            interactive=args.interactive,
        )

        if not rows:
            raise RuntimeError(
                "No annotation rows were generated from the supplied input."
            )

        if args.format in {"csv", "both"}:
            write_csv(rows, csv_path)

        if args.format in {"xlsx", "both"}:
            write_xlsx(rows, xlsx_path)

        print_summary(summary, output_paths)
        return 0

    except (
        FileNotFoundError,
        FileExistsError,
        RuntimeError,
        ValueError,
        OSError,
    ) as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
